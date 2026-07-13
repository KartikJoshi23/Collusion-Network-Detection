"""
Load the 10 TechNova Excel files into DuckDB.

Idempotent: safe to run repeatedly — `CREATE OR REPLACE TABLE` wipes and re-inserts.

Called from Makefile via `make ingest` (inside the backend container).
Can also be imported: `from app.data.duckdb_loader import load_all_excels, get_connection`
"""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path

import duckdb
import pandas as pd

from app.config import get_settings

logger = logging.getLogger(__name__)

# Each Excel filename in `data/structured` maps to a DuckDB table name.
# Strip the "NN_" prefix and lowercase. "01_Departments.xlsx" -> "departments".
_FILE_TO_TABLE_OVERRIDES: dict[str, str] = {
    "05_Products_Services.xlsx": "products_services",
    "08_Financial_Transactions.xlsx": "financial_transactions",
    "09_Training_Compliance.xlsx": "training_compliance",
    "10_Assets_Licenses.xlsx": "assets_licenses",
}


def _derive_table_name(filename: str) -> str:
    if filename in _FILE_TO_TABLE_OVERRIDES:
        return _FILE_TO_TABLE_OVERRIDES[filename]
    # "01_Departments.xlsx" -> "departments"
    stem = Path(filename).stem
    without_prefix = re.sub(r"^\d+_", "", stem)
    return without_prefix.lower()


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Lowercase columns, strip whitespace, replace spaces/dashes with underscores."""
    df = df.copy()
    df.columns = [
        re.sub(r"[\s\-]+", "_", str(c).strip().lower()) for c in df.columns
    ]
    return df


def _pick_data_sheet(path: Path) -> str:
    """Each TechNova workbook has a 'Schema_Notes' sheet plus the real data sheet.
    Pick the first sheet whose name isn't 'Schema_Notes'."""
    import openpyxl  # lazy import — only needed for introspection

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        if name.strip().lower() != "schema_notes":
            return name
    # Fallback: if there's only one sheet, use it even if it's called Schema_Notes.
    return wb.sheetnames[0]


def get_connection(read_only: bool = False) -> duckdb.DuckDBPyConnection:
    """Return a DuckDB connection to the persisted TechNova DB.

    Creates the parent directory if missing.
    """
    settings = get_settings()
    db_path = Path(settings.duckdb_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    return duckdb.connect(str(db_path), read_only=read_only)


def load_all_excels() -> dict[str, int]:
    """Load every `.xlsx` file in structured_data_dir into DuckDB.

    Returns a dict of {table_name: row_count}.
    Skips the README/Schema workbook.
    """
    settings = get_settings()
    source_dir = Path(settings.structured_data_dir)
    if not source_dir.exists():
        raise FileNotFoundError(f"Structured data directory not found: {source_dir}")

    excel_files = sorted(p for p in source_dir.glob("*.xlsx") if "README" not in p.name.upper())
    if not excel_files:
        raise FileNotFoundError(f"No data .xlsx files in {source_dir}")

    loaded: dict[str, int] = {}
    con = get_connection(read_only=False)
    try:
        for path in excel_files:
            table_name = _derive_table_name(path.name)
            sheet_name = _pick_data_sheet(path)
            logger.info("Loading %s[%s] -> table '%s'", path.name, sheet_name, table_name)
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            df = _normalise_columns(df)

            # Register the DataFrame with DuckDB, then CTAS so it's persisted.
            con.register("staging_df", df)
            con.execute(f'CREATE OR REPLACE TABLE "{table_name}" AS SELECT * FROM staging_df')
            con.unregister("staging_df")

            row_count = con.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            loaded[table_name] = row_count
            logger.info("  -> %s rows", row_count)
    finally:
        con.close()

    return loaded


def list_tables() -> list[str]:
    con = get_connection(read_only=True)
    try:
        rows = con.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main' ORDER BY table_name"
        ).fetchall()
    finally:
        con.close()
    return [r[0] for r in rows]


def main() -> None:
    logging.basicConfig(
        level=os.getenv("LOG_LEVEL", "INFO"),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    counts = load_all_excels()
    print("Loaded tables:")
    for name, n in sorted(counts.items()):
        print(f"  {name:<28} {n:>5} rows")


if __name__ == "__main__":
    main()
